# 🔐 Authentication & Authorization
from django.contrib.auth import (
    authenticate, login as auth_login, logout, update_session_auth_hash
)
from django.contrib.auth.decorators import login_required

# 🖼 View Rendering & Redirect
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string

# 🧠 ORM & Query Utilities
from django.db.models import Q, Min, Max, Count, Exists, OuterRef

# 📬 Email Support
from django.core.mail import send_mail, EmailMessage

# ⚙️ Django Settings & Utility
from django.conf import settings
from django.utils import timezone
from django.utils.timezone import now
from django.urls import reverse

# 📤 HTTP Responses
from django.http import HttpResponse, FileResponse

# 🧾 PDF Generation
from xhtml2pdf import pisa

# 🧮 Excel File Generation
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage



import openpyxl
from django.http import HttpResponse


# 🧰 Python Built-ins
import os
import io
import base64
import random
import string

# ✅ Messages & Models
from django.contrib import messages
from .models import *
from education.models import *
from accounts.models import *

# 🌐 Pagination
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

# 📲 Custom Utilities for send message 
from admission.utils import *
from dotenv import load_dotenv
# ✅ Load environment variables
load_dotenv()


# সেশন ডাটা  ধরে রাখে
from django.urls import reverse

from io import BytesIO




from django.shortcuts import redirect
from django.http import HttpResponse
import pdfkit


def Admission_Notice(request):
    now = timezone.now()

    # সক্রিয় ইনটেকগুলো খুঁজে বের করা
    active_intakes = Intake.objects.filter(is_open=True).order_by('start_date')

    # শর্ত অনুযায়ী সক্রিয় ইনটেক নির্বাচন
    if active_intakes.exists():
        # একাধিক ইনটেক থাকলে
        ongoing_intakes = [i for i in active_intakes if i.start_date <= now <= i.end_date]
        upcoming_intakes = [i for i in active_intakes if i.start_date > now]

        if ongoing_intakes:
            # যদি কোনো ইনটেক চলমান থাকে
            intake = ongoing_intakes[-1]  # সর্বশেষ শুরু হওয়া ইনটেক
        elif upcoming_intakes:
            # যদি কোনো ইনটেক ভবিষ্যতে শুরু হয়
            intake = upcoming_intakes[0]
        else:
            # যদি কোনো ইনটেক চলমান না থাকে
            intake = active_intakes.last()
    else:
        intake = None

    # নোটিশগুলো ফিল্টার করা
    notices = AdmissionNotice.objects.filter(active=True).order_by('-create_at')[:10]
    notice_id = request.GET.get('notice_id')
    selected_notice = None
    if notice_id:
        selected_notice = AdmissionNotice.objects.filter(id=notice_id, active=True).first()
    elif notices.exists():
        selected_notice = notices.first()

    context = {
        'notices': notices,
        'selected_notice': selected_notice,
        'admission_details': intake,
    }
    return render(request, 'admission/admission_notices.html', context)





def syllabus_info(request):

    syllabus = Course.objects.all().order_by('position_no') 

    context = {
 
        'syllabus': syllabus
    }

    return render(request, 'admission/syllabus.html', context)




def admission_procedure(request):

    return render(request, 'admission/procedure.html')




@login_required
def personal_info(request):
    try:
        # এখানে user এর মাধ্যমে UserRegistration মডেলটি খুঁজে বের করলাম
        applicant = UserRegistration.objects.get(email=request.user.email)
    except UserRegistration.DoesNotExist:
        messages.error(request, "You need to complete applicant registration first!")
        return redirect("home")  # Redirect to home or applicant registration page

    # আগের ডাটা থাকলে সেটি নিয়ে আসবে, না থাকলে None
    personal_info = PersonalInformation.objects.filter(applicant=applicant).first()

    applied_intake = Intake.objects.filter(
    is_open=True,
    start_date__lte=timezone.now(),
    end_date__gte=timezone.now()
).order_by("-start_date").first()

    if request.method == "GET":
        messages.info(request, "Please fill out your profile to complete the application.")
    
    # To show dynamic data from model to template
    personal_choices = PersonalInformation.get_personal_all_choices()

    if request.method == "POST":
        # ইনপুট ফিল্ড থেকে ডাটা গ্রহণ
        father_name = request.POST.get("ap_f_name")  
        mother_name = request.POST.get("ap_m_name")  
        date_of_birth = request.POST.get("ap_dob")  
        sex = request.POST.get("ap_sex")  
        religion = request.POST.get("ap_religion")  
        blood_group = request.POST.get("ap_bg")  
        nationality = request.POST.get("ap_cntry")  
        national_id = request.POST.get("ap_nid")  
        present_address = request.POST.get("ap_prsnt_add")  
        permanent_address = request.POST.get("ap_prmnt_add")  

        # আগের ডাটা থাকলে আপডেট করবে, না থাকলে নতুন এন্ট্রি তৈরি করবে
        personal_info, created = PersonalInformation.objects.update_or_create(
            applicant=applicant,
            defaults={
                "father_name": father_name,
                "mother_name": mother_name,
                "date_of_birth": date_of_birth,
                "sex": sex,
                "religion": religion,
                "blood_group": blood_group,
                "nationality": nationality,
                "national_id": national_id,
                "present_address": present_address,
                "permanent_address": permanent_address,
            }
        )

        messages.success(request, "Your personal information has been saved successfully!")
        return redirect(reverse("academic_qualification"))



    context = {
        "applied_intake": applied_intake,
        "applicant": applicant,
        "personal_info": personal_info,
        "personal_choices": personal_choices,
        # 
        "active_step": "personal",
    }
    return render(request, "admission/personal_info.html", context)




@login_required
def academic_qualification(request):
    try:
        applicant = UserRegistration.objects.get(email=request.user.email)
    except UserRegistration.DoesNotExist:
        messages.error(request, "You need to complete applicant registration first!")
        return redirect("home")

    academic_info = AcademicQualification.objects.filter(applicant=applicant).first()
    choices = AcademicQualification.get_all_choices()

    if request.method == "POST":
        print(request.POST)

        def get_post_value(field_name):
            value = request.POST.get(field_name)
            return value if value not in ["", None] else None

        data = {
            "ssc_exam": get_post_value("ssc_exam"),
            "ssc_board": get_post_value("ssc_board"),
            "ssc_roll": get_post_value("ssc_roll"),
            "ssc_result_type": get_post_value("ssc_result_type"),
            "ssc_result": get_post_value("ssc_result"),
            "ssc_group": get_post_value("ssc_group"),
            "ssc_passing_year": get_post_value("ssc_passing_year"),
            "hsc_exam": get_post_value("hsc_exam"),
            "hsc_board": get_post_value("hsc_board"),
            "hsc_roll": get_post_value("hsc_roll"),
            "hsc_result_type": get_post_value("hsc_result_type"),
            "hsc_result": get_post_value("hsc_result"),
            "hsc_group": get_post_value("hsc_group"),
            "hsc_passing_year": get_post_value("hsc_passing_year"),
            "graduation_exam": get_post_value("graduation_exam"),
            "graduation_subject": get_post_value("graduation_subject"),
            "graduation_university": get_post_value("graduation_university"),
            "graduation_result_type": get_post_value("graduation_result_type"),
            "graduation_result": get_post_value("graduation_result"),
            "graduation_passing_year": get_post_value("graduation_passing_year"),
            "graduation_course_duration": get_post_value("graduation_course_duration"),
            "post_graduation_exam": get_post_value("post_graduation_exam"),
            "post_graduation_subject": get_post_value("post_graduation_subject"),
            "post_graduation_university": get_post_value("post_graduation_university"),
            "post_graduation_result_type": get_post_value("post_graduation_result_type"),
            "post_graduation_result": get_post_value("post_graduation_result"),
            "post_graduation_passing_year": get_post_value("post_graduation_passing_year"),
            "post_graduation_course_duration": get_post_value("post_graduation_course_duration"),
        }

        # Handle file fields and delete old files if new ones are uploaded
        if request.FILES.get("graduation_certificate"):
            if academic_info and academic_info.graduation_certificate:
                academic_info.graduation_certificate.delete(save=False)
            data["graduation_certificate"] = request.FILES["graduation_certificate"]

        if request.FILES.get("graduation_transcript"):
            if academic_info and academic_info.graduation_transcript:
                academic_info.graduation_transcript.delete(save=False)
            data["graduation_transcript"] = request.FILES["graduation_transcript"]

        academic_info, created = AcademicQualification.objects.update_or_create(
            applicant=applicant,
            defaults=data
        )

        messages.success(request, "Your academic qualification information has been saved successfully!")
        return redirect(reverse("application_review"))

    context = {
        "applicant": applicant,
        "academic_info": academic_info,
        "choices": choices,
        "active_step": "academic",
    }
    return render(request, "admission/academic_qualification.html", context)



@login_required
def application_review(request):
    try:
        applicant = UserRegistration.objects.get(email=request.user.email)
    except UserRegistration.DoesNotExist:
        messages.error(request, "You need to complete applicant registration first!")
        return redirect("home")

    personal_info = PersonalInformation.objects.filter(applicant=applicant).first()
    academic_info = AcademicQualification.objects.filter(applicant=applicant).first()
    applicant_photo = Photo.objects.filter(applicant=applicant).first()

    if request.method == "POST":
        image_uploaded = "applicant_image" in request.FILES
        sig_uploaded = "applicant_sig" in request.FILES

        # ফটো অবজেক্ট না থাকলে তৈরি করি
        if not applicant_photo:
            applicant_photo = Photo(applicant=applicant)

        # ফাইল এলে পুরাতন মুছে আপডেট করি
        if image_uploaded:
            if applicant_photo.applicant_image:
                applicant_photo.applicant_image.delete(save=False)
            applicant_photo.applicant_image = request.FILES["applicant_image"]

        if sig_uploaded:
            if applicant_photo.applicant_sig:
                applicant_photo.applicant_sig.delete(save=False)
            applicant_photo.applicant_sig = request.FILES["applicant_sig"]

        # আপডেট করার পর এখন যাচাই করি: আছে কি না
        has_photo = applicant_photo.applicant_image
        has_signature = applicant_photo.applicant_sig

        if has_photo and has_signature:
            applicant_photo.save()
            messages.success(request, "Your application has been submitted successfully. Kindly proceed to complete your payment.")
            return redirect("payment")
        else:
            messages.error(request, "Please upload both your photo and signature before submitting.")
            return redirect("application_review")

    context = {
        "applicant": applicant,
        "personal_info": personal_info,
        "academic_info": academic_info,
        "application": applicant_photo,
        "active_step": "review",
    }
    return render(request, "admission/application_review.html", context)

def payment(request):

    context ={
        "active_step": "payment",
    }
    return render(request, 'admission/payment.html', context)





# local host
def admit_card_pdf(request):
    if not request.user.is_authenticated:
        messages.warning(request, "You need to log in first.")
        return redirect('login')

    if request.user.role != 'applicant':
        messages.warning(request, "You are not authorized to access this page.")
        return redirect('home')

    applicant = request.user

    try:
        if applicant.payment.status != 'Paid':
            messages.warning(request, "You must complete payment before downloading the admit card.")
            return redirect('home')
    except Exception:
        messages.warning(request, "Payment record not found. Please complete payment first.")
        return redirect('home')

    # Applicant's photo and signature
    photo_uri = ''
    sig_uri = ''
    if hasattr(applicant, 'applicant_photo') and applicant.applicant_photo:
        if applicant.applicant_photo.applicant_image:
            photo_uri = request.build_absolute_uri(applicant.applicant_photo.applicant_image.url)
        if applicant.applicant_photo.applicant_sig:
            sig_uri = request.build_absolute_uri(applicant.applicant_photo.applicant_sig.url)

    # Intake and coordinator's signature
    intake = applicant.intake
    coordinator_signature_uri = ''
    if intake and intake.coordinator_signature:
        coordinator_signature_uri = request.build_absolute_uri(intake.coordinator_signature.url)

    context = {
        "applicant": applicant,
        "photo_uri": photo_uri,
        "sig_uri": sig_uri,
        "ju_logo_uri": request.build_absolute_uri(settings.STATIC_URL + 'img/ju_logo.png'),
        "coordinator_signature_uri": coordinator_signature_uri,
        "personal_info": getattr(applicant, 'personal_info', None),
        "intake": intake,
        "now": timezone.now(),
    }

    html_string = render_to_string("admission/admit_card_download.html", context)

    options = {
        'page-size': 'A4',
        'encoding': "UTF-8",
        'enable-local-file-access': None,
        'margin-top': '5mm',
        'margin-bottom': '5mm',
    }

    config = pdfkit.configuration(
        wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    )

    pdf = pdfkit.from_string(html_string, False, options=options, configuration=config)

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="admit_card.pdf"'

    return response

# সার্ভারের জন্য নিচের কোড



# import os
# from django.conf import settings
# from django.http import HttpResponse
# from django.template.loader import render_to_string
# from django.utils import timezone
# from django.contrib import messages
# from django.shortcuts import redirect
# import pdfkit


# def admit_card_pdf(request):
#     if not request.user.is_authenticated:
#         messages.warning(request, "You need to log in first.")
#         return redirect('login')

#     if request.user.role != 'applicant':
#         messages.warning(request, "You are not authorized to access this page.")
#         return redirect('home')

#     applicant = request.user

#     try:
#         if applicant.payment.status != 'Paid':
#             messages.warning(request, "You must complete payment before downloading the admit card.")
#             return redirect('home')
#     except Exception:
#         messages.warning(request, "Payment record not found. Please complete payment first.")
#         return redirect('home')

#     # Paths for local image access using file://
#     photo_uri = ''
#     sig_uri = ''
#     if hasattr(applicant, 'applicant_photo') and applicant.applicant_photo:
#         if applicant.applicant_photo.applicant_image:
#             photo_path = os.path.join(settings.MEDIA_ROOT, applicant.applicant_photo.applicant_image.name)
#             photo_uri = f'file://{photo_path}'
#         if applicant.applicant_photo.applicant_sig:
#             sig_path = os.path.join(settings.MEDIA_ROOT, applicant.applicant_photo.applicant_sig.name)
#             sig_uri = f'file://{sig_path}'

#     intake = applicant.intake
#     coordinator_signature_uri = ''
#     if intake and intake.coordinator_signature:
#         coord_sig_path = os.path.join(settings.MEDIA_ROOT, intake.coordinator_signature.name)
#         coordinator_signature_uri = f'file://{coord_sig_path}'

#     ju_logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'JU_logo.png')
#     ju_logo_uri = f'file://{ju_logo_path}'


#     context = {
#         "applicant": applicant,
#         "photo_uri": photo_uri,
#         "sig_uri": sig_uri,
#         "ju_logo_uri": ju_logo_uri,
#         "coordinator_signature_uri": coordinator_signature_uri,
#         "personal_info": getattr(applicant, 'personal_info', None),
#         "intake": intake,
#         "now": timezone.now(),
#     }

#     html_string = render_to_string("admission/admit_card_download.html", context)

#     options = {
#         'page-size': 'A4',
#         'encoding': "UTF-8",
#         'enable-local-file-access': '',  # must be empty string, not None or True
#         'margin-top': '5mm',
#         'margin-bottom': '5mm',
#     }

#     config = pdfkit.configuration(wkhtmltopdf='/usr/bin/wkhtmltopdf')

#     try:
#         pdf = pdfkit.from_string(html_string, False, options=options, configuration=config)
#     except Exception as e:
#         return HttpResponse(f"PDF generation failed: {e}", status=500)

#     response = HttpResponse(pdf, content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="admit_card.pdf"'
#     return response




def image_resize(request):
    return render(request, 'admission/image_resize.html')


# ----------------------- --------------

def applicant_status(request):
    if not request.user.is_authenticated or request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to view this page.")
        return redirect('home')  

    # সব ব্যাচ
    batches = Intake.objects.values_list('batch_name', flat=True).distinct()
    latest_intake = Intake.objects.order_by('-create_at').first()
    default_batch = latest_intake.batch_name if latest_intake else None

    # ইউজার সিলেক্ট করেছে কিনা?
    selected_batch = request.GET.get('batch_name') or default_batch
    show_details = request.GET.get('details') == '1'

    # ব্যাচ অনুযায়ী রেজিস্ট্রেশন করা এপ্লিকেন্টস
    applicants = UserRegistration.objects.filter(
        role='applicant',
        intake__batch_name=selected_batch
    ).select_related('payment')

    total_registered = applicants.count()

    # প্রোফাইল complete কিনা চেক করতে annotated queryset
    applicants = applicants.annotate(
        has_personal_info=Exists(PersonalInformation.objects.filter(applicant=OuterRef('pk'))),
        has_academic_info=Exists(AcademicQualification.objects.filter(applicant=OuterRef('pk'))),
        has_photo=Exists(Photo.objects.filter(applicant=OuterRef('pk')))
    )

    # প্রোফাইল completed
    profile_completed_qs = applicants.filter(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    profile_completed = profile_completed_qs.count()
    profile_incomplete = total_registered - profile_completed

    # পেমেন্ট paid/pending শুধু প্রোফাইল completed দের মধ্য থেকে
    payment_done_qs = profile_completed_qs.filter(payment__status='Paid')
    payment_pending_qs = profile_completed_qs.filter(Q(payment__status='Pending') | Q(payment__isnull=True))

    # বিকাশ/নগদ এর হিসাব
    bkash_paid_qs = payment_done_qs.filter(payment__payment_method='bkash')
    nagad_paid_qs = payment_done_qs.filter(payment__payment_method='nagad')

    context = {
        'batches': batches,
        'selected_batch': selected_batch,
        'total_registered': total_registered,
        'profile_completed': profile_completed,
        'profile_incomplete': profile_incomplete,
        'payment_done': payment_done_qs.count(),
        'payment_pending': payment_pending_qs.count(),
        'bkash_count': bkash_paid_qs.count(),
        'nagad_count': nagad_paid_qs.count(),
        'show_details': show_details,

        # শুধুমাত্র details অন করলে list পাঠাবে
        'bkash_paid': bkash_paid_qs if show_details else [],
        'nagad_paid': nagad_paid_qs if show_details else [],
        'profile_pending': applicants.exclude(pk__in=profile_completed_qs).all() if show_details else [],
    }

    return render(request, 'manager/applicant_status.html', context)



def applicant_list(request):
    if not request.user.is_authenticated or request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to view this page.")
        return redirect('home')

    batches = Intake.objects.values_list('batch_name', flat=True).distinct()
    latest_intake = Intake.objects.order_by('-create_at').first()
    default_batch = latest_intake.batch_name if latest_intake else None

    selected_batch = request.GET.get('batch_name') or default_batch
    filter_type = request.GET.get('filter')
    applicant_list_query = request.GET.get('applicant_list_query', '').strip()

    applicants = UserRegistration.objects.filter(role=UserRegistration.APPLICANT)
    
    if selected_batch:
        applicants = applicants.filter(intake__batch_name=selected_batch)

    # ✅ Corrected search filter using Q()
    if applicant_list_query:
        applicants = applicants.filter(
            Q(full_name__icontains=applicant_list_query) |
            Q(mobile_number__icontains=applicant_list_query) |
            Q(roll_number__icontains=applicant_list_query)
        )

    applicants = applicants.annotate(
        has_personal_info=Exists(PersonalInformation.objects.filter(applicant=OuterRef('pk'))),
        has_academic_info=Exists(AcademicQualification.objects.filter(applicant=OuterRef('pk'))),
        has_photo=Exists(Photo.objects.filter(applicant=OuterRef('pk')))
    )

    profile_completed = applicants.filter(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    profile_incomplete = applicants.exclude(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    paid_applicants = applicants.filter(payment__status='Paid')
    pending_applicants = profile_completed.filter(Q(payment__status='Pending') | Q(payment__isnull=True))
    incomplete_applicants = profile_incomplete.filter(Q(payment__status='Pending') | Q(payment__isnull=True))

    if filter_type == 'paid':
        display_applicants = paid_applicants
    elif filter_type == 'pending':
        display_applicants = pending_applicants
    elif filter_type == 'incomplete':
        display_applicants = incomplete_applicants
    else:
        display_applicants = applicants

    total_paid = paid_applicants.count()
    total_pending = pending_applicants.count()
    total_incomplete = incomplete_applicants.count()
    total_all = applicants.count()

    # ✅ Pagination
    paginator = Paginator(display_applicants.select_related('payment', 'applicant_photo'), 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)

    context = {
        'batches': batches,
        'selected_batch': selected_batch,
        'filter_type': filter_type,
        'page_obj': page_obj,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'total_incomplete': total_incomplete,
        'total_all': total_all,
        'applicant_list_query': applicant_list_query,
    }

    return render(request, 'manager/applicant_list.html', context)


def export_applicants_excel(request):
    if not request.user.is_authenticated or request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to view this page.")
        return redirect('home')

    selected_batch = request.GET.get('batch_name')
    filter_type = request.GET.get('filter')  # paid, pending, incomplete, all
    applicant_list_query = request.GET.get('applicant_list_query', '').strip()

    applicants = UserRegistration.objects.filter(role=UserRegistration.APPLICANT)

    # Filter by batch
    if selected_batch:
        applicants = applicants.filter(intake__batch_name=selected_batch)

    # Filter by search query (applicant_list_query)
    if applicant_list_query:
        applicants = applicants.filter(
            Q(full_name__icontains=applicant_list_query) |
            Q(mobile_number__icontains=applicant_list_query) |
            Q(roll_number__icontains=applicant_list_query)
        )

    # Annotate profile completion
    applicants = applicants.annotate(
        has_personal_info=Exists(PersonalInformation.objects.filter(applicant=OuterRef('pk'))),
        has_academic_info=Exists(AcademicQualification.objects.filter(applicant=OuterRef('pk'))),
        has_photo=Exists(Photo.objects.filter(applicant=OuterRef('pk')))
    )

    profile_completed = applicants.filter(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    profile_incomplete = applicants.exclude(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    # Apply filter_type
    if filter_type == 'paid':
        applicants = applicants.filter(payment__status='Paid')
    elif filter_type == 'pending':
        applicants = profile_completed.filter(Q(payment__status='Pending') | Q(payment__isnull=True))
    elif filter_type == 'incomplete':
        applicants = profile_incomplete.filter(Q(payment__status='Pending') | Q(payment__isnull=True))
    # else: keep all

    # Now export filtered applicants
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applicants"

    headers = [
        "Name", "Email", "Phone", "Batch", "roll_number",
        "Father Name", "Mother Name", "DOB", "Sex", "Religion", "Blood Group", "Nationality", "National ID",
        "Present Address", "Permanent Address",

        # SSC
        "SSC Exam", "SSC Board", "SSC Roll", "SSC Result Type", "SSC Result", "SSC Group", "SSC Passing Year",

        # HSC
        "HSC Exam", "HSC Board", "HSC Roll", "HSC Result Type", "HSC Result", "HSC Group", "HSC Passing Year",

        # Graduation
        "Grad Exam", "Grad Subject", "Grad University", "Grad Result Type", "Grad Result", "Grad Passing Year", "Grad Duration",

        # Post-Graduation
        "PG Exam", "PG Subject", "PG University", "PG Result Type", "PG Result", "PG Passing Year", "PG Duration",

        # Payment
        "Transaction ID", "Amount", "Payment Status", "Method",

        # Images
        "Photo", "Signature"
    ]
    ws.append(headers)

    for app in applicants:
        personal = getattr(app, 'personal_info', None)
        academic = getattr(app, 'academic_qualification', None)
        payment = getattr(app, 'payment', None)
        photo = getattr(app, 'applicant_photo', None)

        row = [
            app.full_name, app.email, app.mobile_number, app.intake.batch_name if app.intake else '', app.roll_number,

            # Personal
            getattr(personal, 'father_name', ''), getattr(personal, 'mother_name', ''), getattr(personal, 'date_of_birth', ''),
            getattr(personal, 'sex', ''), getattr(personal, 'religion', ''), getattr(personal, 'blood_group', ''),
            getattr(personal, 'nationality', ''), getattr(personal, 'national_id', ''),
            getattr(personal, 'present_address', ''), getattr(personal, 'permanent_address', ''),

            # SSC
            getattr(academic, 'ssc_exam', ''), getattr(academic, 'ssc_board', ''), getattr(academic, 'ssc_roll', ''),
            getattr(academic, 'ssc_result_type', ''), getattr(academic, 'ssc_result', ''), getattr(academic, 'ssc_group', ''),
            getattr(academic, 'ssc_passing_year', ''),

            # HSC
            getattr(academic, 'hsc_exam', ''), getattr(academic, 'hsc_board', ''), getattr(academic, 'hsc_roll', ''),
            getattr(academic, 'hsc_result_type', ''), getattr(academic, 'hsc_result', ''), getattr(academic, 'hsc_group', ''),
            getattr(academic, 'hsc_passing_year', ''),

            # Graduation
            getattr(academic, 'graduation_exam', ''), getattr(academic, 'graduation_subject', ''),
            getattr(academic, 'graduation_university', ''), getattr(academic, 'graduation_result_type', ''),
            getattr(academic, 'graduation_result', ''), getattr(academic, 'graduation_passing_year', ''),
            getattr(academic, 'graduation_course_duration', ''),

            # Post Graduation
            getattr(academic, 'post_graduation_exam', ''), getattr(academic, 'post_graduation_subject', ''),
            getattr(academic, 'post_graduation_university', ''), getattr(academic, 'post_graduation_result_type', ''),
            getattr(academic, 'post_graduation_result', ''), getattr(academic, 'post_graduation_passing_year', ''),
            getattr(academic, 'post_graduation_course_duration', ''),

            # Payment
            getattr(payment, 'transaction_id', ''), getattr(payment, 'amount', ''),
            getattr(payment, 'status', ''), getattr(payment, 'payment_method', ''),

            '', ''  # placeholders for images
        ]
        ws.append(row)

        current_row = ws.max_row

        # Photo
        if photo and photo.applicant_image:
            img_path = os.path.join(settings.MEDIA_ROOT, photo.applicant_image.name)
            if os.path.exists(img_path):
                img = ExcelImage(img_path)
                img.width = 60
                img.height = 60
                ws.add_image(img, f"{get_column_letter(len(headers)-1)}{current_row}")

        # Signature
        if photo and photo.applicant_sig:
            sig_path = os.path.join(settings.MEDIA_ROOT, photo.applicant_sig.name)
            if os.path.exists(sig_path):
                sig = ExcelImage(sig_path)
                sig.width = 80
                sig.height = 20
                ws.add_image(sig, f"{get_column_letter(len(headers))}{current_row}")

    # Auto adjust column width
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in [get_column_letter(len(headers)-1), get_column_letter(len(headers))]:
            continue
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Create downloadable Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"applicants_{selected_batch or 'all'}_{filter_type or 'all'}_{applicant_list_query or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def applicant_detail(request, applicant_id):
    if not request.user.is_authenticated or request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to view this page.")
        return redirect('home')
        
    applicant = get_object_or_404(UserRegistration, id=applicant_id, role=UserRegistration.APPLICANT)

    personal_info = getattr(applicant, 'personal_info', None)
    academic_info = getattr(applicant, 'academic_qualification', None)
    payment = getattr(applicant, 'payment', None)
    photo = getattr(applicant, 'applicant_photo', None)

    context = {
        'applicant': applicant,
        'personal_info': personal_info,
        'academic_info': academic_info,
        'payment': payment,
        'photo': photo,
    }

    return render(request, 'manager/applicant_detail.html', context)




def send_message_to_applicants(request):
    if not request.user.is_authenticated or request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to view this page.")
        return redirect('home')

    # ✅ Check balance
    balance_response = get_sms_balance()
    balance = balance_response.get('balance', 'Unavailable')
    sms_limit = balance_response.get('sms_limit', 'N/A')

    # ✅ Intake logic
    batches = Intake.objects.values_list('batch_name', flat=True).distinct()
    latest_intake = Intake.objects.order_by('-create_at').first()
    default_batch = latest_intake.batch_name if latest_intake else None
    selected_batch = request.GET.get('batch_name') or default_batch

    selected_intake = Intake.objects.filter(batch_name=selected_batch).order_by('-create_at').first()
    if not selected_intake:
        messages.warning(request, "No intake found for the selected batch.")
        return redirect('home')

    applicants = UserRegistration.objects.filter(
        role=UserRegistration.APPLICANT,
        intake=selected_intake
    ).select_related('payment')

    # Profile completeness annotations
    applicants = applicants.annotate(
        has_personal_info=Exists(PersonalInformation.objects.filter(applicant=OuterRef('pk'))),
        has_academic_info=Exists(AcademicQualification.objects.filter(applicant=OuterRef('pk'))),
        has_photo=Exists(Photo.objects.filter(applicant=OuterRef('pk')))
    )

    profile_completed_qs = applicants.filter(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    profile_incomplete_qs = applicants.exclude(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    paid_applicants = applicants.filter(payment__status='Paid')

    pending_applicants = profile_completed_qs.filter(
        Q(payment__status='Pending') | Q(payment__isnull=True)
    )

    incomplete_applicants = profile_incomplete_qs.filter(
        Q(payment__status='Pending') | Q(payment__isnull=True)
    )

    filter_type = request.GET.get('filter')
    show_paid = filter_type == 'paid'
    show_pending = filter_type == 'pending'
    show_incomplete = filter_type == 'incomplete'

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_applicants')
        message_text = request.POST.get('message')

        selected_applicants = UserRegistration.objects.filter(id__in=selected_ids)

        success_count = 0
        failed = []

        for applicant in selected_applicants:
            raw_number = applicant.mobile_number
            number = normalize_number(raw_number)

            result = send_sms(number, message_text)


            if result.get("status") == "SUCCESS":
                success_count += 1
            else:
                failed.append(number)


        messages.success(request, f"Message successfully sent to {success_count} applicant(s).")

        if failed:
            messages.warning(request, f"Failed to deliver message to{', '.join(failed)}")

        return redirect('send_message_to_applicants')

    context = {
        'batches': batches,
        'selected_batch': selected_batch,
        'latest_intake': selected_intake,
        'paid_applicants': paid_applicants if show_paid else [],
        'pending_applicants': pending_applicants if show_pending else [],
        'incomplete_applicants': incomplete_applicants if show_incomplete else [],
        'filter_type': filter_type,
        'balance': balance,
        'sms_limit': sms_limit,
    }

    return render(request, 'manager/send_sms.html', context)



def send_email_to_applicants(request):
    if not request.user.is_authenticated or request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to view this page.")
        return redirect('home')

    # Batch list and default selection
    batches = Intake.objects.values_list('batch_name', flat=True).distinct()
    latest_intake = Intake.objects.order_by('-create_at').first()
    default_batch = latest_intake.batch_name if latest_intake else None
    selected_batch = request.GET.get('batch_name') or default_batch

    # Intake filter by selected batch
    selected_intake = Intake.objects.filter(batch_name=selected_batch).order_by('-create_at').first()
    if not selected_intake:
        messages.warning(request, "No intake found for the selected batch.")
        return redirect('home')

    applicants = UserRegistration.objects.filter(role='applicant', intake=selected_intake).select_related('payment')

    # Annotate with profile completeness status
    applicants = applicants.annotate(
        has_personal_info=Exists(PersonalInformation.objects.filter(applicant=OuterRef('pk'))),
        has_academic_info=Exists(AcademicQualification.objects.filter(applicant=OuterRef('pk'))),
        has_photo=Exists(Photo.objects.filter(applicant=OuterRef('pk')))
    )

    profile_completed_qs = applicants.filter(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    profile_incomplete_qs = applicants.exclude(
        has_personal_info=True,
        has_academic_info=True,
        has_photo=True
    )

    paid_applicants = applicants.filter(payment__status='Paid')

    pending_applicants = profile_completed_qs.filter(
        Q(payment__status='Pending') | Q(payment__isnull=True)
    )

    incomplete_applicants = profile_incomplete_qs.filter(
        Q(payment__status='Pending') | Q(payment__isnull=True)
    )

    # Determine filter
    filter_type = request.GET.get('filter')
    show_paid = filter_type == 'paid'
    show_pending = filter_type == 'pending'
    show_incomplete = filter_type == 'incomplete'

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_applicants')
        email_subject = request.POST.get('subject')
        email_message = request.POST.get('message')
        files = request.FILES.getlist('attachments')

        selected_applicants = UserRegistration.objects.filter(id__in=selected_ids)
        recipient_list = [applicant.email for applicant in selected_applicants]

        email = EmailMessage(
            subject=email_subject,
            body=email_message,
            from_email=settings.EMAIL_HOST_USER,
            to=recipient_list,
        )

        for file in files:
            email.attach(file.name, file.read(), file.content_type)

        email.send(fail_silently=False)

        messages.success(
            request,
            f"📧 Email sent to {len(recipient_list)} applicant(s) with {len(files)} attachment(s)."
        )
        return redirect('send_email_to_applicants')

    context = {
        'batches': batches,
        'selected_batch': selected_batch,
        'latest_intake': selected_intake,
        'paid_applicants': paid_applicants if show_paid else [],
        'pending_applicants': pending_applicants if show_pending else [],
        'incomplete_applicants': incomplete_applicants if show_incomplete else [],
        'filter_type': filter_type,
    }

    return render(request, 'manager/send_email.html', context)



@login_required
def written_attendance_pdf(request):
    if request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to access this page.")
        return redirect('home')

    intakes = Intake.objects.all().order_by('-create_at')
    selected_intake_id = request.GET.get('intake_id')
    selected_intake = Intake.objects.filter(id=selected_intake_id).first() if selected_intake_id else intakes.first()

    start_roll = request.GET.get('start_roll')
    end_roll = request.GET.get('end_roll')

    applicants = UserRegistration.objects.filter(
        intake=selected_intake,
        payment__status='Paid',
        role=UserRegistration.APPLICANT
    ).order_by('roll_number')

    if start_roll and end_roll:
        applicants = applicants.filter(
            roll_number__gte=start_roll,
            roll_number__lte=end_roll
        )

    applicant_data = []
    for idx, app in enumerate(applicants, start=1):
        photo = getattr(app, 'applicant_photo', None)
        personal = getattr(app, 'personal_info', None)
        academic = getattr(app, 'academic_qualification', None)

        image_data_uri = ''
        if photo and photo.applicant_image:
            image_path = os.path.join(settings.MEDIA_ROOT, str(photo.applicant_image))
            if os.path.exists(image_path):
                with open(image_path, 'rb') as image_file:
                    encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                    image_data_uri = f"data:image/jpeg;base64,{encoded_string}"

        applicant_data.append({
            'sl': idx,
            'roll_number': app.roll_number,
            'applicant': app,            
            'personal': personal,        
            'academic': academic,        
            'image_data_uri': image_data_uri,
        })

    # JU Logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img/ju_logo.png')
    ju_logo_uri = ''
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as logo_file:
            encoded_logo = base64.b64encode(logo_file.read()).decode('utf-8')
            ju_logo_uri = f"data:image/png;base64,{encoded_logo}"

    context = {
        'intakes': intakes,
        'selected_intake': selected_intake,
        'applicants': applicant_data,
        'start_roll': start_roll,
        'end_roll': end_roll,
        'ju_logo_uri': ju_logo_uri,
    }

    if 'generate_pdf' in request.GET:
        html_string = render_to_string('manager/written_attendance.html', context)
        pdf_buffer = io.BytesIO()
        pisa_status = pisa.CreatePDF(src=html_string, dest=pdf_buffer)

        if pisa_status.err:
            return HttpResponse('PDF generation error: ' + str(pisa_status.err))

        pdf_buffer.seek(0)
        return FileResponse(pdf_buffer, as_attachment=True, filename='applicants_with_images.pdf')

    return render(request, 'manager/written_attendance_form.html', context)

# for servar



from django.contrib.staticfiles import finders

# @login_required
# def written_attendance_pdf(request):
#     if request.user.role != UserRegistration.MANAGER:
#         messages.warning(request, "You are not authorized to access this page.")
#         return redirect('home')

#     intakes = Intake.objects.all().order_by('-create_at')
#     selected_intake_id = request.GET.get('intake_id')
#     selected_intake = Intake.objects.filter(id=selected_intake_id).first() if selected_intake_id else intakes.first()

#     start_roll = request.GET.get('start_roll')
#     end_roll = request.GET.get('end_roll')

#     applicants = UserRegistration.objects.filter(
#         intake=selected_intake,
#         payment__status='Paid',
#         role=UserRegistration.APPLICANT
#     ).order_by('roll_number')

#     if start_roll and end_roll:
#         applicants = applicants.filter(
#             roll_number__gte=start_roll,
#             roll_number__lte=end_roll
#         )

#     applicant_data = []
#     for idx, app in enumerate(applicants, start=1):
#         photo = getattr(app, 'applicant_photo', None)
#         personal = getattr(app, 'personal_info', None)
#         academic = getattr(app, 'academic_qualification', None)

#         image_data_uri = ''
#         if photo and photo.applicant_image:
#             image_path = os.path.join(settings.MEDIA_ROOT, str(photo.applicant_image))
#             if os.path.exists(image_path):
#                 with open(image_path, 'rb') as image_file:
#                     encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
#                     image_data_uri = f"data:image/jpeg;base64,{encoded_string}"

#         applicant_data.append({
#             'sl': idx,
#             'roll_number': app.roll_number,
#             'applicant': app,
#             'personal': personal,
#             'academic': academic,
#             'image_data_uri': image_data_uri,
#         })

#     # ✅ Static থেকে JU Logo reliably লোড করা
#     ju_logo_uri = ''
#     logo_path = finders.find('img/JU_logo.png')

#     if logo_path and os.path.exists(logo_path):
#         with open(logo_path, 'rb') as logo_file:
#             encoded_logo = base64.b64encode(logo_file.read()).decode('utf-8')
#             ju_logo_uri = f"data:image/png;base64,{encoded_logo}"

#     context = {
#         'intakes': intakes,
#         'selected_intake': selected_intake,
#         'applicants': applicant_data,
#         'start_roll': start_roll,
#         'end_roll': end_roll,
#         'ju_logo_uri': ju_logo_uri,
#     }

#     if 'generate_pdf' in request.GET:
#         html_string = render_to_string('manager/written_attendance.html', context)
#         pdf_buffer = io.BytesIO()
#         pisa_status = pisa.CreatePDF(src=html_string, dest=pdf_buffer)

#         if pisa_status.err:
#             return HttpResponse('PDF generation error: ' + str(pisa_status.err))

#         pdf_buffer.seek(0)
#         return FileResponse(pdf_buffer, as_attachment=True, filename='applicants_with_images.pdf')

#     return render(request, 'manager/written_attendance_form.html', context)

@login_required
def viva_attendance_pdf(request):
    if request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "You are not authorized to access this page.")
        return redirect('home')

    intakes = Intake.objects.all().order_by('-create_at')
    selected_intake_id = request.GET.get('intake_id')
    selected_intake = Intake.objects.filter(id=selected_intake_id).first() if selected_intake_id else intakes.first()

    start_roll = request.GET.get('start_roll')
    end_roll = request.GET.get('end_roll')

    applicants = UserRegistration.objects.filter(
        intake=selected_intake,
        payment__status='Paid',
        role=UserRegistration.APPLICANT
    ).order_by('roll_number')

    if start_roll and end_roll:
        applicants = applicants.filter(
            roll_number__gte=start_roll,
            roll_number__lte=end_roll
        )

    applicant_data = []
    for idx, app in enumerate(applicants, start=1):
        photo = getattr(app, 'applicant_photo', None)
        personal = getattr(app, 'personal_info', None)
        academic = getattr(app, 'academic_qualification', None)

        image_data_uri = ''
        if photo and photo.applicant_image:
            image_path = os.path.join(settings.MEDIA_ROOT, str(photo.applicant_image))
            if os.path.exists(image_path):
                with open(image_path, 'rb') as image_file:
                    encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                    image_data_uri = f"data:image/jpeg;base64,{encoded_string}"

        applicant_data.append({
            'sl': idx,
            'roll_number': app.roll_number,
            'applicant': app,           
            'personal': personal,        
            'academic': academic,       
            'image_data_uri': image_data_uri,
        })

    # JU Logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img/ju_logo.png')
    ju_logo_uri = ''
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as logo_file:
            encoded_logo = base64.b64encode(logo_file.read()).decode('utf-8')
            ju_logo_uri = f"data:image/png;base64,{encoded_logo}"

    context = {
        'intakes': intakes,
        'selected_intake': selected_intake,
        'applicants': applicant_data,
        'start_roll': start_roll,
        'end_roll': end_roll,
        'ju_logo_uri': ju_logo_uri,
    }

    if 'generate_pdf' in request.GET:
        html_string = render_to_string('manager/viva_attendance.html', context)
        pdf_buffer = io.BytesIO()
        pisa_status = pisa.CreatePDF(src=html_string, dest=pdf_buffer)

        if pisa_status.err:
            return HttpResponse('PDF generation error: ' + str(pisa_status.err))

        pdf_buffer.seek(0)
        return FileResponse(pdf_buffer, as_attachment=True, filename='applicants_with_images.pdf')

    return render(request, 'manager/viva_attendance_form.html', context)



@login_required(login_url='login')  
def result_prepare(request):
    if not hasattr(request.user, 'role') or request.user.role != UserRegistration.MANAGER:
        messages.warning(request, "⛔ You are not authorized to access this page.")
        return redirect('home')

    # Get all batches for filtering
    batches = Intake.objects.filter(is_open=True).order_by('-start_date')

    # Default to the latest batch
    default_batch = batches.first() if batches else None

    # Get the batch filter from GET parameters (if any)
    selected_batch = request.GET.get('batch', default_batch.batch_name if default_batch else None)

    # Filter applicants: roll_number is set AND payment is Paid
    applicants = UserRegistration.objects.select_related('result_prepare', 'payment').filter(
        roll_number__isnull=False,
        intake__batch_name=selected_batch,
        payment__status='Paid'
    ).order_by('result_prepare__category', '-result_prepare__total_marks')

    # Download Excel if 'download_excel' in GET params
    if 'download_excel' in request.GET:

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Exam Results"

        headers = ['SL', 'Roll', 'Name', 'Email', 'Mobile', 'Category', 'MCQ Marks', 'Job Marks', 'Academic Marks', 'Total Marks']
        sheet.append(headers)

        sl = 1
        for applicant in applicants:
            result = getattr(applicant, 'result_prepare', None)
            sheet.append([
                sl,
                applicant.roll_number,
                applicant.full_name,
                applicant.email,
                applicant.mobile_number,
                result.category if result else '',
                result.mcq_marks if result else '',
                result.job_marks if result else '',
                result.academic_marks if result else '',
                result.total_marks if result else '',
            ])
            sl += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=exam_results.xlsx'
        workbook.save(response)
        return response

    # POST: Save individual applicant's marks
    if request.method == 'POST':
        applicant_id = request.POST.get('save_id')
        applicant = get_object_or_404(UserRegistration, id=applicant_id)

        category = request.POST.get(f'category_{applicant_id}')
        mcq_input = request.POST.get(f'mcq_{applicant_id}', '').strip()
        job_input = request.POST.get(f'job_{applicant_id}', '').strip()

        try:
            mcq_marks = float(mcq_input) if mcq_input else 0.0
        except ValueError:
            mcq_marks = 0.0

        try:
            job_marks = float(job_input) if job_input else 0.0
        except ValueError:
            job_marks = 0.0

        result, created = ResultPrepare.objects.get_or_create(applicant=applicant)
        result.category = category
        result.mcq_marks = mcq_marks
        result.job_marks = job_marks
        result.save()

        messages.success(request, f"✅ {applicant.full_name} এর জন্য মার্কস সফলভাবে সংরক্ষণ করা হয়েছে।")

        # Retain selected batch in redirect URL
        selected_batch_post = request.POST.get('selected_batch', selected_batch)
        return redirect(f'{reverse("result_prepare")}?batch={selected_batch_post}')

    return render(request, 'manager/result_prepare.html', {
        'applicants': applicants,
        'batches': batches,
        'selected_batch': selected_batch
    })






